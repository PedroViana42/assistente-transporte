from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from app.importers.column_mapper import NORMALIZED_COLUMN_ALIASES, _normalize_column_name


ExcelSheets = dict[str, pd.DataFrame]


def list_sheet_names(file_path: str | Path) -> list[str]:
    path = _validate_xlsx_path(file_path)

    excel_file = pd.ExcelFile(path, engine="openpyxl")
    return excel_file.sheet_names


def read_excel_sheets(file_path: str | Path) -> ExcelSheets:
    path = _validate_xlsx_path(file_path)

    sheet_names = _list_sheets_with_order_number(path)
    if not sheet_names:
        return {}

    sheets = pd.read_excel(path, sheet_name=sheet_names, engine="openpyxl")
    if isinstance(sheets, pd.DataFrame):
        sheets = {sheet_names[0]: sheets}

    cleaned_sheets: ExcelSheets = {}

    for sheet_name, dataframe in sheets.items():
        cleaned_dataframe = _remove_empty_rows(dataframe)
        if cleaned_dataframe.empty:
            continue

        cleaned_sheets[sheet_name] = cleaned_dataframe

    return cleaned_sheets


def _validate_xlsx_path(file_path: str | Path) -> Path:
    path = Path(file_path)

    if path.suffix.lower() != ".xlsx":
        raise ValueError("O arquivo deve ter extensao .xlsx.")

    if not path.is_file():
        raise FileNotFoundError(f"Arquivo nao encontrado: {path}")

    return path


def _remove_empty_rows(dataframe: pd.DataFrame) -> pd.DataFrame:
    return dataframe.dropna(how="all").reset_index(drop=True)


def _list_sheets_with_order_number(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_names = []
        for worksheet in workbook.worksheets:
            first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if first_row is None:
                continue

            canonical_columns = {
                NORMALIZED_COLUMN_ALIASES.get(_normalize_column_name(str(cell)))
                for cell in first_row
                if cell is not None
            }
            if "order_number" in canonical_columns:
                sheet_names.append(worksheet.title)

        return sheet_names
    finally:
        workbook.close()
